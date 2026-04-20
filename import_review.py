"""
审核队列导入脚本
读取业务人员填写完的 review_queue.xlsx
将"审核结论"列写回数据库

用法: python import_review.py [review_queue.xlsx]
"""
import sys, sqlite3, time
sys.path.insert(0, 'C:/Users/jiawa/skc_color')

import openpyxl
from database import DB_PATH

FAMILY_NAMES = {
    0:'白/米白', 1:'黄', 2:'橙', 3:'红',
    4:'粉',     5:'紫', 6:'蓝', 7:'绿', 8:'棕', 9:'灰/黑'
}


def import_review(xlsx_path: str, reviewer: str = 'import_script'):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 找列索引
    header = [str(c.value or '').strip() for c in ws[1]]
    try:
        idx_pantone    = header.index('潘通色号') + 1
        idx_conclusion = header.index('审核结论') + 1
        idx_skc        = header.index('当前SKC') + 1
    except ValueError as e:
        print('ERROR: 找不到必要列 -', e)
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    now  = time.strftime('%Y-%m-%d %H:%M:%S')

    confirmed = skipped = invalid = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        pantone    = str(row[idx_pantone - 1] or '').strip().upper()
        conclusion = str(row[idx_conclusion - 1] or '').strip()
        skc        = str(row[idx_skc - 1] or '').strip()

        if not pantone:
            continue

        # 结论为空 → 跳过
        if not conclusion:
            skipped += 1
            continue

        # 解析结论：支持 "1"、"1 黄"、"黄" 等格式
        final_fam = None
        for k, v in FAMILY_NAMES.items():
            if conclusion == str(k) or conclusion.startswith(str(k)+' ') or conclusion == v:
                final_fam = k
                break

        if final_fam is None:
            print('  WARN: 无法解析结论 "%s" (潘通:%s)，跳过' % (conclusion, pantone))
            invalid += 1
            continue

        # 更新主表
        conn.execute(
            "UPDATE skc_colors SET family=?, status='active' WHERE pantone=?",
            (final_fam, pantone)
        )
        # 更新审核队列
        conn.execute(
            "UPDATE review_queue SET status='confirmed', reviewer=?, final_skc=?, reviewed_at=? WHERE pantone=?",
            (reviewer, skc, now, pantone)
        )
        confirmed += 1
        print('  OK: %s → 色系%d(%s)' % (pantone, final_fam, FAMILY_NAMES[final_fam]))

    conn.commit()
    conn.close()

    print()
    print('导入完成: confirmed=%d  skipped=%d  invalid=%d' % (confirmed, skipped, invalid))

    # 剩余待审核数
    conn2 = sqlite3.connect(DB_PATH)
    remaining = conn2.execute(
        "SELECT COUNT(*) FROM review_queue WHERE status='pending'"
    ).fetchone()[0]
    conn2.close()
    print('剩余待审核: %d' % remaining)


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'C:/Users/jiawa/skc_color/review_queue.xlsx'
    rev  = sys.argv[2] if len(sys.argv) > 2 else 'import_script'
    import_review(path, rev)
