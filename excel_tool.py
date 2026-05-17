"""
Excel 工具：读取潘通号列 → 自动填充SKC编码
用法：python excel_tool.py input.xlsx [潘通号列名] [输出文件名]
默认列名：pantone
"""
import sys, csv, re
sys.path.insert(0, 'C:/Users/jiawa/skc_color')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from classifier import classify_from_hex
from shade_mapper import shade_to_range
from batch_encoder import map_shade_base

# 加载潘通对照表（同时索引 TPG 和 base，便于 TCX 输入时 fallback）
PANTONE_DB = {}
PANTONE_BASE_DB = {}  # key=pantone_base（不含后缀），value=首条记录
with open('C:/Users/jiawa/skc_color/pantone_skc_v2.csv', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        key = r['pantone'].strip().upper()
        PANTONE_DB[key] = r
        # 提取 base（去掉 TPG/TCX 后缀）
        base = key
        for sfx in ('TPG', 'TCX'):
            if base.endswith(sfx):
                base = base[:-len(sfx)]
                break
        if base not in PANTONE_BASE_DB:
            PANTONE_BASE_DB[base] = r

FAMILY_NAMES = {
    '0':'白/米白','1':'黄','2':'橙','3':'红',
    '4':'粉','5':'紫','6':'蓝','7':'绿','8':'棕','9':'灰/黑'
}

# 色系对应填充色（用于色块预览）
FAMILY_PREVIEW_COLOR = {
    '0':'FFF8F0','1':'FFF176','2':'FFB74D','3':'EF5350',
    '4':'F48FB1','5':'CE93D8','6':'64B5F6','7':'81C784',
    '8':'A1887F','9':'B0BEC5'
}


def hex_to_openpyxl(hex_color: str) -> str:
    return hex_color.lstrip('#').upper()


def lookup_pantone(pantone_input: str) -> dict | None:
    """
    查找潘通号，支持多种格式和系列：
    - 19-4150TPG / 19-4150TCX / 19-4150 均可命中
    """
    key = pantone_input.strip().upper()
    # 1. 精确匹配
    if key in PANTONE_DB:
        return PANTONE_DB[key]
    # 2. 补全 TPG 后缀
    if not key.endswith('TPG') and not key.endswith('TCX'):
        if key + 'TPG' in PANTONE_DB:
            return PANTONE_DB[key + 'TPG']
    # 3. TCX → 查 base（因为现有数据是 TPG）
    base = key
    for sfx in ('TPG', 'TCX'):
        if base.endswith(sfx):
            base = base[:-len(sfx)]
            break
    if base in PANTONE_BASE_DB:
        return PANTONE_BASE_DB[base]
    return None


def process_excel(input_path: str, pantone_col: str = 'pantone', output_path: str = None):
    if output_path is None:
        output_path = input_path.replace('.xlsx', '_skc.xlsx')

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 找到潘通号列索引
    header_row = [cell.value for cell in ws[1]]
    try:
        pantone_idx = next(i for i, h in enumerate(header_row)
                          if h and pantone_col.lower() in str(h).lower()) + 1
    except StopIteration:
        print('ERROR: 找不到列 "%s"，现有列：%s' % (pantone_col, header_row))
        return

    # 在末尾追加新列
    last_col = ws.max_column
    new_cols = {
        'SKC编码':    last_col + 1,
        '色系':       last_col + 2,
        '色阶':       last_col + 3,
        '色阶区间':   last_col + 4,
        '颜色预览':   last_col + 5,
        '需审核':     last_col + 6,
        'TPG色号':    last_col + 7,
        'TCX色号':    last_col + 8,
    }

    # 写表头
    header_fill = PatternFill('solid', fgColor='2C3E50')
    header_font = Font(color='FFFFFF', bold=True)
    for col_name, col_idx in new_cols.items():
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 处理每行
    found = not_found = review = 0
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in range(2, ws.max_row + 1):
        pantone_cell = ws.cell(row=row_idx, column=pantone_idx)
        pantone_val = pantone_cell.value
        if not pantone_val:
            continue

        result = lookup_pantone(str(pantone_val))

        if result:
            skc          = result['skc']
            family       = result['family']
            shade        = result['shade']
            shade_range  = result['shade_range']
            hex_color    = result['hex']
            needs_review = result['needs_review'] == 'True'
            # 互转色号：从原始 pantone_base 推算
            p_base = result['pantone'].upper()
            for sfx in ('TPG', 'TCX'):
                if p_base.endswith(sfx):
                    p_base = p_base[:-len(sfx)]
                    break
            tpg_val = p_base + 'TPG'
            tcx_val = p_base + 'TCX'
            found += 1
            if needs_review:
                review += 1
        else:
            skc = shade = shade_range = '未找到'
            family = '-'
            hex_color = None
            needs_review = False
            tpg_val = tcx_val = '未找到'
            not_found += 1

        # 写入各列
        ws.cell(row=row_idx, column=new_cols['SKC编码'],  value=skc)
        ws.cell(row=row_idx, column=new_cols['色系'],     value=FAMILY_NAMES.get(family, family))
        ws.cell(row=row_idx, column=new_cols['色阶'],     value=shade)
        ws.cell(row=row_idx, column=new_cols['色阶区间'], value=shade_range)
        ws.cell(row=row_idx, column=new_cols['需审核'],   value='是' if needs_review else '否')
        ws.cell(row=row_idx, column=new_cols['TPG色号'],  value=tpg_val)
        ws.cell(row=row_idx, column=new_cols['TCX色号'],  value=tcx_val)

        # 颜色预览色块
        preview_cell = ws.cell(row=row_idx, column=new_cols['颜色预览'], value='  ')
        if hex_color:
            preview_cell.fill = PatternFill('solid', fgColor=hex_to_openpyxl(hex_color))

        # SKC单元格：需审核标红
        skc_cell = ws.cell(row=row_idx, column=new_cols['SKC编码'])
        if needs_review:
            skc_cell.fill = PatternFill('solid', fgColor='FFEB3B')
            skc_cell.font = Font(bold=True, color='B71C1C')
        elif skc != '未找到':
            skc_cell.font = Font(bold=True, color='1B5E20')

        # 加边框
        for col_idx in new_cols.values():
            ws.cell(row=row_idx, column=col_idx).border = border

    # 调整列宽
    col_widths = {'SKC编码':10,'色系':10,'色阶':8,'色阶区间':18,'颜色预览':10,'需审核':8,'TPG色号':14,'TCX色号':14}
    for col_name, width in col_widths.items():
        ws.column_dimensions[get_column_letter(new_cols[col_name])].width = width

    wb.save(output_path)
    print('完成！找到:%d  未找到:%d  需审核:%d' % (found, not_found, review))
    print('输出文件:', output_path)


if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print('用法: python excel_tool.py input.xlsx [潘通号列名] [输出文件名]')
        sys.exit(1)
    input_path  = args[0]
    pantone_col = args[1] if len(args) > 1 else 'pantone'
    output_path = args[2] if len(args) > 2 else None
    process_excel(input_path, pantone_col, output_path)
