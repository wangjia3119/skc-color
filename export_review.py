"""
导出审核队列为 Excel
供业务人员逐条确认边界色归属
"""
import sys, sqlite3
sys.path.insert(0, 'C:/Users/jiawa/skc_color')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from database import DB_PATH
from classifier import classify_from_hex

FAMILY_NAMES = {
    0:'白/米白', 1:'黄', 2:'橙', 3:'红',
    4:'粉',     5:'紫', 6:'蓝', 7:'绿', 8:'棕', 9:'灰/黑'
}

# 每个色系的候选邻居（审核时展示可能的替代归属）
NEIGHBOR_FAMILIES = {
    1: [2],   # 黄 ↔ 橙
    2: [1, 8],# 橙 ↔ 黄/棕
    3: [4],   # 红 ↔ 粉
    4: [3, 5],# 粉 ↔ 红/紫
    5: [4],   # 紫 ↔ 粉
    6: [7],   # 蓝 ↔ 绿
    7: [6],   # 绿 ↔ 蓝
    8: [2],   # 棕 ↔ 橙
}


def export_review_queue(output_path: str = 'C:/Users/jiawa/skc_color/review_queue.xlsx'):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # 取审核队列 + 主表信息
    rows = conn.execute("""
        SELECT q.pantone, q.name, q.hex, q.candidate_1, q.status,
               s.family, s.shade, s.shade_range, s.lab_l, s.lab_a, s.lab_b, s.skc
        FROM review_queue q
        LEFT JOIN skc_colors s ON q.pantone = s.pantone
        WHERE q.status = 'pending'
        ORDER BY s.family, s.lab_l DESC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '待审核颜色'

    # 样式定义
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center')
    yellow_fill = PatternFill('solid', fgColor='FFF9C4')
    green_fill  = PatternFill('solid', fgColor='E8F5E9')

    # 表头
    headers = [
        '序号', '潘通色号', '颜色名称', '当前SKC', '当前色系',
        '颜色预览', 'L*', 'a*', 'b*', '色阶', '色阶区间',
        '建议归属1', '建议归属2', '审核结论', '备注'
    ]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 数据行
    for seq, row in enumerate(rows, 1):
        family = row['family'] if row['family'] is not None else -1
        hex_val = row['hex'] or ''
        lab_l = round(row['lab_l'], 1) if row['lab_l'] else ''
        lab_a = round(row['lab_a'], 1) if row['lab_a'] else ''
        lab_b = round(row['lab_b'], 1) if row['lab_b'] else ''

        # 计算候选归属
        neighbors = NEIGHBOR_FAMILIES.get(family, [])
        suggest1 = '%d %s' % (family, FAMILY_NAMES.get(family,'?')) if family != -1 else '?'
        suggest2 = '%d %s' % (neighbors[0], FAMILY_NAMES.get(neighbors[0],'?')) if neighbors else '—'

        data = [
            seq,
            row['pantone'],
            row['name'],
            row['skc'] or row['candidate_1'],
            '%d %s' % (family, FAMILY_NAMES.get(family,'?')) if family != -1 else '?',
            '',           # 颜色预览（色块）
            lab_l, lab_a, lab_b,
            row['shade'],
            row['shade_range'],
            suggest1,
            suggest2,
            '',           # 审核结论（业务填写）
            '',           # 备注
        ]
        ws.append(data)
        r = ws.max_row

        # 颜色预览色块
        if hex_val:
            preview = ws.cell(row=r, column=6)
            preview.fill = PatternFill('solid', fgColor=hex_val.lstrip('#').upper())
            preview.value = '  '

        # 行底色交替
        row_fill = yellow_fill if seq % 2 == 0 else green_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            cell.alignment = center
            if col_idx not in (6, 14, 15):  # 预览列和填写列不覆盖
                cell.fill = row_fill

        # 审核结论列：下拉提示（加底色区分）
        conclusion_cell = ws.cell(row=r, column=14)
        conclusion_cell.fill = PatternFill('solid', fgColor='FFFDE7')

    # 列宽
    col_widths = [6,16,22,10,12,10,7,7,7,8,20,14,14,14,20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行高
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # 统计 sheet
    ws2 = wb.create_sheet('统计')
    ws2.append(['色系', '待审核数', '占比'])
    from collections import Counter
    dist = Counter(str(row['family']) for row in rows)
    total = len(rows)
    for k in sorted(dist, key=lambda x: int(x) if x and x != 'None' else -1):
        name = FAMILY_NAMES.get(int(k) if k and k != 'None' else -1, '?')
        count = dist[k]
        ws2.append(['%s %s' % (k, name), count, '%.1f%%' % (count/total*100)])
    ws2.append(['合计', total, '100%'])

    wb.save(output_path)
    print('审核队列已导出: %s' % output_path)
    print('待审核总数: %d' % total)


if __name__ == '__main__':
    export_review_queue()
